import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.formatting.rule import FormulaRule

def load_and_clean_data(file):
    if file.name.endswith('.csv'):
        df = pd.read_csv(file, header=None)
    else:
        df = pd.read_excel(file, header=None)
        
    header_row = df[df.apply(lambda row: row.astype(str).str.replace(" ", "").str.contains('일자').any(), axis=1)].index
    if len(header_row) > 0:
        df.columns = df.iloc[header_row[0]]
        df = df.iloc[header_row[0]+1:].reset_index(drop=True)
        
    df.columns = [str(c).replace(" ", "") for c in df.columns]
    
    if '일자' in df.columns and '요일' in df.columns:
        df['일자'] = df['일자'].ffill()
        df['요일'] = df['요일'].ffill()
        
    df['일자_dt'] = pd.to_datetime(df['일자'], errors='coerce')
    df = df.dropna(subset=['일자_dt'])
    df['week_start'] = df['일자_dt'] - pd.to_timedelta(df['일자_dt'].dt.weekday, unit='d')
    return df

def shorten_subject(name):
    if pd.isna(name) or not str(name).strip(): 
        return ""
    name_str = str(name).strip()
    
    mapping = {
        "간호관리": "관리", "기본간호1": "기본1", "기본간호2": "기본2",
        "기초약리": "약리", "기초영양": "영양", "기초치과": "치과", 
        "기초한방": "한방", "기초해부": "해부", "노인간호": "노인", 
        "모성간호": "모성", "모자보건": "모자", "보건교육": "보교", 
        "보건행정": "보행", "산업보건": "산업", "성인간호1": "성인1", 
        "성인간호2": "성인2", "아동간호": "아동", "응급간호": "응급", 
        "의료관계법규": "법규", "의학용어": "의용", "지역사회": "지역",
        "질병관리사업": "질병", "인구와출산": "인구", "환경보건": "환경"
    }
    for k, v in mapping.items():
        if k in name_str: return v
        
    keywords = ["관리", "기본1", "기본2", "약리", "영양", "치과", "한방", "해부", "노인", "모성", "모자", "보교", "보행", "산업", "성인1", "성인2", "아동", "응급", "법규", "의용", "지역", "인구", "질병", "환경"]
    for kw in keywords:
        if kw in name_str: return kw
        
    return name_str

def create_attendance_excel(df, target_weeks, student_names):
    student_count = len(student_names)
    wb = Workbook()
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") 
    stat_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") 
    cumul_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") 
    light_grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") 
    
    dropout_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    dropout_font = Font(color="808080", strike=True)
    
    title_font = Font(size=14, bold=True, color="1F4E78")
    bold_font = Font(bold=True)
    subject_font = Font(size=10) 
    teacher_font = Font(size=9, color="595959")
    
    stat_title_font = Font(size=10, bold=True)
    stat_head_font = Font(size=9, bold=True)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False) 
    shrink_align = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)
    
    sheets = {
        "통합입력": wb.active,
        "기초간호": wb.create_sheet("기초간호"),
        "보건간호": wb.create_sheet("보건간호"),
        "공중간호": wb.create_sheet("공중간호")
    }
    wb.active.title = "통합입력"
    subject_map = {"기초간호": "기초간호", "보건간호": "보건간호", "공중간호": "공중보건", "통합입력": "통합"}
    
    for ws in sheets.values():
        ws.sheet_view.showGridLines = False 
        ws.freeze_panes = "B1" 
        
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE 
        ws.page_setup.paperSize = ws.PAPERSIZE_A4 
        ws.page_setup.fitToPage = True 
        ws.page_setup.fitToWidth = 1 
        ws.page_setup.fitToHeight = 0 
        
        ws.page_margins.left = 0.2
        ws.page_margins.right = 0.2
        ws.page_margins.top = 0.3
        ws.page_margins.bottom = 0.3
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True
        
        ws.column_dimensions['A'].width = 12 
        for c in range(2, 37): 
            ws.column_dimensions[get_column_letter(c)].width = 3.6 
            
        for c in range(37, 42): 
            ws.column_dimensions[get_column_letter(c)].width = 5.0 
        for c in range(42, 47): 
            ws.column_dimensions[get_column_letter(c)].width = 6.0 
            
    block_height = student_count + 8 
    student_row_height = max(15, min(28, int(630 / student_count)))
    
    for w_idx, current_week_start in enumerate(target_weeks):
        start_row = 1 + (w_idx * block_height)
        title_str = f"■ {current_week_start.strftime('%Y년 %m월 %d일')} 시작 주간"
        
        week_total_classes = {"통합입력": 0, "기초간호": 0, "보건간호": 0, "공중간호": 0}
        schedule_struct = []
        
        dates_in_week = [current_week_start + pd.Timedelta(days=i) for i in range(5)]
        days_kr = ["월", "화", "수", "목", "금"]
        
        for d_idx, current_date in enumerate(dates_in_week):
            day_str = days_kr[d_idx]
            df_day = df[df['일자_dt'].dt.date == current_date]
            day_data = []
            
            cancel_day = ""
            if not df_day.empty:
                first_row = df_day.iloc[0]
                if str(first_row.get('요일', '')) not in ['월','화','수','목','금'] and len(str(first_row.get('요일', ''))) > 1:
                    cancel_day = str(first_row.get('요일', ''))
                    
            for p in range(1, 8):
                main_subj, sub_subj, teacher_name, cancel_reason = "", "", "", cancel_day
                is_class = False
                
                if not cancel_day and not df_day.empty:
                    p_row = df_day[df_day['교시'].astype(str).str.contains(str(p))]
                    if not p_row.empty:
                        row_data = p_row.iloc[0]
                        if '휴강' in str(row_data.get('과목코드','')) or '휴강' in str(row_data.get('세부교과','')):
                            cancel_reason = "휴강"
                        else:
                            main_subj = str(row_data.get('교과목', '')).strip()
                            raw_sub = str(row_data.get('세부교과', '')).strip()
                            raw_teacher = str(row_data.get('훈련교사', '')).strip()
                            if raw_teacher and raw_teacher != 'nan':
                                teacher_name = raw_teacher
                                
                            if raw_sub and raw_sub != 'nan':
                                is_class = True
                                sub_subj = shorten_subject(raw_sub)
                
                day_data.append({'p': p, 'main': main_subj, 'sub': sub_subj, 'teacher': teacher_name, 'cancel': cancel_reason, 'is_class': is_class})
                
                for sheet_name in week_total_classes.keys():
                    target_keyword = subject_map[sheet_name]
                    is_my_subject = (target_keyword == "통합") or (target_keyword in main_subj)
                    if not cancel_reason and is_class and is_my_subject:
                        week_total_classes[sheet_name] += 1
                        
            schedule_struct.append({'date': current_date, 'day_str': day_str, 'periods': day_data})

        for ws in sheets.values():
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=10)
            ws.cell(row=start_row, column=1, value=title_str).font = title_font
            ws.row_dimensions[start_row].height = 25
            
            c = ws.cell(row=start_row+1, column=1, value="학생이름")
            ws.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+4, end_column=1)
            c.alignment = center_align; c.fill = header_fill; c.border = thin_border
            
            ws.row_dimensions[start_row+1].height = 22 # 요일
            ws.row_dimensions[start_row+2].height = 18 # 교시
            ws.row_dimensions[start_row+3].height = 23 # 과목명
            ws.row_dimensions[start_row+4].height = 18 # 교사명
            
            for i in range(student_count):
                r = start_row + 5 + i
                
                if ws.title == "통합입력":
                    if w_idx == 0:
                        ws.cell(row=r, column=1, value=student_names[i]).alignment = shrink_align
                    else:
                        ws.cell(row=r, column=1, value=f"=A{r - block_height}").alignment = shrink_align
                else:
                    ws.cell(row=r, column=1, value=f"='통합입력'!A{r}").alignment = shrink_align
                    
                ws.cell(row=r, column=1).border = thin_border
                ws.row_dimensions[r].height = student_row_height 
                
        col_idx = 2
        
        for day_info in schedule_struct:
            header_str = f"{day_info['date'].strftime('%m/%d')}({day_info['day_str']})"
            
            for ws in sheets.values():
                ws.merge_cells(start_row=start_row+1, start_column=col_idx, end_row=start_row+1, end_column=col_idx+6)
                c = ws.cell(row=start_row+1, column=col_idx, value=header_str)
                c.alignment = center_align; c.fill = header_fill; c.font = bold_font; c.border = thin_border
                
            for sheet_name, ws in sheets.items():
                target_keyword = subject_map[sheet_name]
                disp_list = []
                
                for p_info in day_info['periods']:
                    is_my_subject = (target_keyword == "통합") or (target_keyword in p_info['main'])
                    if p_info['cancel']:
                        disp_list.append((f"[{p_info['cancel']}]", ""))
                    elif not p_info['is_class'] or not is_my_subject:
                        disp_list.append(("", "")) 
                    else:
                        disp_list.append((p_info['sub'], p_info['teacher']))
                
                start_p = 0
                while start_p < 7:
                    curr_item = disp_list[start_p]
                    end_p = start_p
                    while end_p + 1 < 7 and disp_list[end_p + 1] == curr_item:
                        end_p += 1
                        
                    curr_subj, curr_teacher = curr_item
                    c_start = col_idx + start_p
                    c_end = col_idx + end_p
                    
                    scell = ws.cell(row=start_row+3, column=c_start, value=curr_subj)
                    scell.font = subject_font
                    scell.alignment = shrink_align
                    
                    tcell = ws.cell(row=start_row+4, column=c_start, value=curr_teacher)
                    tcell.font = teacher_font
                    tcell.alignment = shrink_align
                    
                    if start_p != end_p:
                        ws.merge_cells(start_row=start_row+3, start_column=c_start, end_row=start_row+3, end_column=c_end)
                        ws.merge_cells(start_row=start_row+4, start_column=c_start, end_row=start_row+4, end_column=c_end)
                        
                    for p_idx in range(start_p, end_p + 1):
                        act_c = col_idx + p_idx
                        ws.cell(row=start_row+2, column=act_c, value=f"{p_idx+1}").alignment = center_align
                        
                        if not curr_subj or curr_subj.startswith("["):
                            ws.cell(row=start_row+3, column=act_c).fill = light_grey_fill
                            ws.cell(row=start_row+4, column=act_c).fill = light_grey_fill
                            for r in range(start_row+5, start_row+5+student_count):
                                ws.cell(row=r, column=act_c).fill = light_grey_fill
                        else:
                            if sheet_name != "통합입력":
                                col_letter = get_column_letter(act_c)
                                for i in range(student_count):
                                    r = start_row+5+i
                                    ws.cell(row=r, column=act_c, value=f"='통합입력'!{col_letter}{r}").alignment = center_align
                                    
                    start_p = end_p + 1
            col_idx += 7
            
        stat_col = col_idx
        stats = ["총수업", "결석", "지각", "조퇴", "출석"]
        cumuls = ["누계수업", "누계결석", "누계지각", "누계조퇴", "누계출석"]
        
        for sheet_name in ["기초간호", "보건간호", "공중간호"]:
            ws = sheets[sheet_name]
            
            ws.merge_cells(start_row=start_row+1, start_column=stat_col, end_row=start_row+1, end_column=stat_col+4)
            sc = ws.cell(row=start_row+1, column=stat_col, value="주간 통계")
            sc.alignment = shrink_align; sc.fill = stat_fill; sc.font = stat_title_font; sc.border = thin_border
            
            ws.merge_cells(start_row=start_row+1, start_column=stat_col+5, end_row=start_row+1, end_column=stat_col+9)
            cc = ws.cell(row=start_row+1, column=stat_col+5, value="누계 통계")
            cc.alignment = shrink_align; cc.fill = cumul_fill; cc.font = stat_title_font; cc.border = thin_border
            
            for i, (s_name, c_name) in enumerate(zip(stats, cumuls)):
                ws.merge_cells(start_row=start_row+2, start_column=stat_col+i, end_row=start_row+4, end_column=stat_col+i)
                c1 = ws.cell(row=start_row+2, column=stat_col+i, value=s_name)
                c1.alignment = shrink_align; c1.font = stat_head_font
                for rf in range(start_row+2, start_row+5): ws.cell(row=rf, column=stat_col+i).fill = stat_fill
                
                ws.merge_cells(start_row=start_row+2, start_column=stat_col+5+i, end_row=start_row+4, end_column=stat_col+5+i)
                c2 = ws.cell(row=start_row+2, column=stat_col+5+i, value=c_name)
                c2.alignment = shrink_align; c2.font = stat_head_font
                for rf in range(start_row+2, start_row+5): ws.cell(row=rf, column=stat_col+5+i).fill = cumul_fill
            
            for i in range(student_count):
                r = start_row + 5 + i
                data_range = f"B{r}:{get_column_letter(stat_col-1)}{r}"
                
                is_drop = f'OR(RIGHT($A{r},1)="x", RIGHT($A{r},1)="X")'
                
                # 💡 핵심 업데이트: 하이픈(-) 입력 시 총수업 시간에서 자동차감 되도록 MAX 수식 적용
                ws.cell(row=r, column=stat_col, value=f'=IF({is_drop}, 0, MAX(0, {week_total_classes[sheet_name]} - COUNTIF({data_range}, "-")))').alignment = shrink_align
                ws.cell(row=r, column=stat_col+1, value=f'=IF({is_drop}, 0, COUNTIF({data_range}, "X"))').alignment = shrink_align
                ws.cell(row=r, column=stat_col+2, value=f'=IF({is_drop}, 0, COUNTIF({data_range}, "지"))').alignment = shrink_align
                ws.cell(row=r, column=stat_col+3, value=f'=IF({is_drop}, 0, COUNTIF({data_range}, "조"))').alignment = shrink_align
                ws.cell(row=r, column=stat_col+4, value=f'=IF({is_drop}, 0, {get_column_letter(stat_col)}{r}-{get_column_letter(stat_col+1)}{r}-{get_column_letter(stat_col+2)}{r}-{get_column_letter(stat_col+3)}{r})').alignment = shrink_align
                
                for j in range(5):
                    current_stat_cell = f"{get_column_letter(stat_col+j)}{r}"
                    if w_idx == 0: 
                        ws.cell(row=r, column=stat_col+5+j, value=f"={current_stat_cell}").alignment = shrink_align
                    else: 
                        prev_cumul_cell = f"{get_column_letter(stat_col+5+j)}{r - block_height}"
                        ws.cell(row=r, column=stat_col+5+j, value=f"={prev_cumul_cell}+{current_stat_cell}").alignment = shrink_align
        
        for ws in sheets.values():
            max_c = stat_col + 9 if ws.title != "통합입력" else stat_col - 1
            for row in range(start_row+1, start_row+5+student_count):
                for col in range(1, max_c + 1):
                    ws.cell(row=row, column=col).border = thin_border

            ws.row_breaks.append(Break(id=start_row + block_height - 1))

    for ws in sheets.values():
        max_c = 46 if ws.title != "통합입력" else 36
        max_r = 1 + (len(target_weeks) * block_height)
        rule = FormulaRule(formula=['OR(RIGHT($A1,1)="x", RIGHT($A1,1)="X")'], fill=dropout_fill, font=dropout_font)
        ws.conditional_formatting.add(f"A1:{get_column_letter(max_c)}{max_r}", rule)

    excel_data = io.BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)
    return excel_data

# --- Streamlit UI ---
st.set_page_config(page_title="스마트 간호학원 출석부", layout="wide")
st.title("📊 스마트 간호학원 출석부 생성기")

uploaded_file = st.file_uploader("1년치 시간표 파일 업로드 (.csv, .xlsx)", type=["csv", "xlsx"])

if uploaded_file is not None:
    df = load_and_clean_data(uploaded_file)
    unique_weeks = sorted(df['week_start'].dropna().dt.date.unique())
    st.success(f"데이터 파싱 완료! 총 {len(unique_weeks)}개의 주간 일정이 확인되었습니다.")
    
    st.markdown("---")
    st.subheader("👨‍🎓 학생 명단 입력")
    st.info("엑셀이나 카톡에 있는 학생 명단을 복사해서 아래 빈칸에 붙여넣기 하세요. (엔터로 구분하여 한 줄에 한 명씩)")
    
    default_names = "\n".join([f"학생{i+1}" for i in range(10)])
    names_input = st.text_area("명단 입력칸:", value=default_names, height=200)
    
    student_names = [name.strip() for name in names_input.split('\n') if name.strip()]
    student_count = len(student_names)
    
    st.write(f"👉 현재 인식된 학생 수: **{student_count}명**")
    st.markdown("---")
    
    if student_count == 0:
        st.warning("학생 이름을 최소 1명 이상 입력해주세요.")
    else:
        option = st.radio("출석부 생성 방식을 선택하세요:", ["전체 주간 한 번에 생성 (1년치 통합 파일)", "특정 주차만 선택해서 생성"])
        
        if option == "전체 주간 한 번에 생성 (1년치 통합 파일)":
            target_weeks_to_process = unique_weeks
            file_name_suffix = "전체일정"
        else:
            selected_week = st.selectbox("생성할 주차를 선택하세요", unique_weeks, format_func=lambda x: f"{x.strftime('%Y년 %m월 %d일')} 시작 주간")
            target_weeks_to_process = [selected_week]
            file_name_suffix = selected_week.strftime('%y%m%d')
            
        if st.button("출석부 엑셀 다운로드"):
            with st.spinner("엑셀 파일을 생성 중입니다..."):
                excel_file = create_attendance_excel(df, target_weeks_to_process, student_names)
                st.download_button(
                    label="📥 완성된 출석부 다운로드",
                    data=excel_file,
                    file_name=f"자동화_출석부_{file_name_suffix}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
